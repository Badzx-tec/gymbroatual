import os
import io
import smtplib
from email.message import EmailMessage
import uuid
import json
import asyncio
import httpx
import binascii
from datetime import datetime, timezone, timedelta
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request, Response, Depends, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel
from typing import Optional, List
from jose import jwt
from passlib.context import CryptContext
from urllib.parse import urlparse
from fido2.server import FIDO2Server
from fido2.webauthn import PublicKeyCredentialRpEntity, AttestationObject, CollectedClientData, AuthenticatorData
from fido2 import cbor

load_dotenv()

MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME")
JWT_SECRET = os.environ.get("JWT_SECRET")
MP_ACCESS_TOKEN = os.environ.get("MERCADOPAGO_ACCESS_TOKEN")
FRONTEND_URL = os.environ.get("FRONTEND_URL", "http://localhost:3000")
BACKEND_PUBLIC_URL = os.environ.get("BACKEND_PUBLIC_URL", "http://localhost:8000")
# SMTP settings (optional). If not set, emails are logged to `email_logs` collection.
SMTP_HOST = os.environ.get("SMTP_HOST")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "0") or 0)
SMTP_USER = os.environ.get("SMTP_USER")
SMTP_PASS = os.environ.get("SMTP_PASS")
SMTP_FROM = os.environ.get("SMTP_FROM", "no-reply@gymbro.local")


def send_email(to_email: str, subject: str, body: str):
    """Send email via SMTP if configured, otherwise store in `email_logs` collection."""
    try:
        if SMTP_HOST and SMTP_PORT:
            msg = EmailMessage()
            msg['Subject'] = subject
            msg['From'] = SMTP_FROM
            msg['To'] = to_email
            msg.set_content(body)
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as s:
                if SMTP_USER and SMTP_PASS:
                    s.starttls()
                    s.login(SMTP_USER, SMTP_PASS)
                s.send_message(msg)
            return True
    except Exception:
        pass
    # fallback: persist in DB for debugging / dev
    try:
        db.email_logs.insert_one({
            "log_id": f"email_{uuid.uuid4().hex[:12]}",
            "to": to_email,
            "subject": subject,
            "body": body,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception:
        pass
    return False
LOGIN_MAX_ATTEMPTS = int(os.environ.get("LOGIN_MAX_ATTEMPTS", "5"))
LOGIN_LOCK_MINUTES = int(os.environ.get("LOGIN_LOCK_MINUTES", "15"))

origins_env = os.environ.get("CORS_ORIGINS", "*")
if origins_env.strip() == "*":
    cors_origins = ["*"]
else:
    cors_origins = [o.strip() for o in origins_env.split(",") if o.strip()]

app = FastAPI(title="GymBro API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=origins_env.strip() != "*",
    allow_methods=["*"],
    allow_headers=["*"],
)

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Setup FIDO2 / WebAuthn server
try:
    rp_url = urlparse(FRONTEND_URL or BACKEND_PUBLIC_URL)
    rp_id = rp_url.hostname or "localhost"
except Exception:
    rp_id = "localhost"
rp = PublicKeyCredentialRpEntity(id=rp_id, name="GymBro")
fido_server = FIDO2Server(rp)

# ============== WEBSOCKET MANAGER ==============

class ConnectionManager:
    def __init__(self):
        self.active: List[WebSocket] = []

    async def connect(self, ws: WebSocket):
        await ws.accept()
        self.active.append(ws)

    def disconnect(self, ws: WebSocket):
        if ws in self.active:
            self.active.remove(ws)

    async def broadcast(self, message: dict):
        dead = []
        for ws in self.active:
            try:
                await ws.send_json(message)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.disconnect(ws)

ws_manager = ConnectionManager()

# ============== MODELS ==============

class UserRegister(BaseModel):
    email: str
    password: str
    name: str

class UserLogin(BaseModel):
    email: str
    password: str

class EmailVerificationRequest(BaseModel):
    email: str

class EmailVerificationConfirm(BaseModel):
    email: str
    code: str

class StudentCreate(BaseModel):
    nome: str
    email: str
    cpf: str
    telefone: Optional[str] = ""
    plano_id: Optional[str] = ""
    tag_rfid: Optional[str] = ""
    biometria_id: Optional[str] = ""
    biometria_template: Optional[str] = ""
    peso: Optional[float] = None
    idade: Optional[int] = None
    altura: Optional[float] = None
    treino: Optional[str] = ""
    dias_presenca: Optional[int] = 0
    status: str = "ativo"
    data_vencimento: Optional[str] = ""
    academy_id: Optional[str] = ""
    peso_kg: Optional[float] = None
    idade: Optional[int] = None
    altura_cm: Optional[float] = None
    treino: Optional[str] = ""
    dias_frequencia: Optional[int] = 0

class StudentUpdate(BaseModel):
    nome: Optional[str] = None
    email: Optional[str] = None
    cpf: Optional[str] = None
    telefone: Optional[str] = None
    plano_id: Optional[str] = None
    tag_rfid: Optional[str] = None
    biometria_id: Optional[str] = None
    biometria_template: Optional[str] = None
    peso: Optional[float] = None
    idade: Optional[int] = None
    altura: Optional[float] = None
    treino: Optional[str] = None
    dias_presenca: Optional[int] = None
    status: Optional[str] = None
    data_vencimento: Optional[str] = None
    academy_id: Optional[str] = None
    peso_kg: Optional[float] = None
    idade: Optional[int] = None
    altura_cm: Optional[float] = None
    treino: Optional[str] = None
    dias_frequencia: Optional[int] = None

class PlanCreate(BaseModel):
    nome: str
    valor: float
    duracao_dias: int
    descricao: Optional[str] = ""
    ativo: bool = True

class PlanUpdate(BaseModel):
    nome: Optional[str] = None
    valor: Optional[float] = None
    duracao_dias: Optional[int] = None
    descricao: Optional[str] = None
    ativo: Optional[bool] = None

class AccessValidation(BaseModel):
    tag_id: str
    tipo: str = "rfid"
    academy_id: Optional[str] = ""

class AcademyCreate(BaseModel):
    nome: str
    endereco: Optional[str] = ""
    telefone: Optional[str] = ""
    cnpj: Optional[str] = ""
    email: Optional[str] = ""
    catraca_ip: Optional[str] = "192.168.1.9"
    catraca_port: Optional[int] = 7878
    ativo: bool = True

class AcademyUpdate(BaseModel):
    nome: Optional[str] = None
    endereco: Optional[str] = None
    telefone: Optional[str] = None
    cnpj: Optional[str] = None
    email: Optional[str] = None
    catraca_ip: Optional[str] = None
    catraca_port: Optional[int] = None
    ativo: Optional[bool] = None

class CatracaCommand(BaseModel):
    action: str  # release_entry, release_exit, block, message
    message: Optional[str] = ""
    academy_id: Optional[str] = ""

class AcademySubscriptionCheckout(BaseModel):
    academy_id: str
    amount: float
    description: Optional[str] = "Mensalidade da academia"
    payer_email: Optional[str] = ""

class CatracaLanExecute(BaseModel):
    academy_id: str
    action: str
    message: Optional[str] = ""
    raw_hex: Optional[str] = ""
    timeout_seconds: Optional[float] = 3.0

class BiometricRegister(BaseModel):
    biometria_id: str

# ============== AUTH HELPERS ==============

def create_token(user_id: str, email: str):
    payload = {
        "user_id": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")

def normalize_email(email: str) -> str:
    return (email or "").strip().lower()

async def academy_payment_is_active(academy_id: str) -> bool:
    if not academy_id:
        return True
    now = datetime.now(timezone.utc)
    docs = await db.academy_billing.find({"academy_id": academy_id, "payment_status": "approved"}, {"_id": 0}).sort("updated_at", -1).to_list(20)
    for doc in docs:
        paid_until = doc.get("paid_until")
        if not paid_until:
            continue
        try:
            dt = datetime.fromisoformat(paid_until)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            if dt >= now:
                return True
        except Exception:
            continue
    return False

def ilnet2_payload_for_action(action: str, message: str = "", raw_hex: str = "") -> bytes:
    if raw_hex:
        cleaned = raw_hex.replace(" ", "")
        try:
            return bytes.fromhex(cleaned)
        except binascii.Error:
            raise HTTPException(status_code=400, detail="raw_hex invalido")

    env_map = {
        "release_entry": os.environ.get("ILNET2_CMD_RELEASE_ENTRY", ""),
        "release_exit": os.environ.get("ILNET2_CMD_RELEASE_EXIT", ""),
        "block": os.environ.get("ILNET2_CMD_BLOCK", ""),
        "message": os.environ.get("ILNET2_CMD_MESSAGE", ""),
    }
    if action in env_map and env_map[action]:
        try:
            return bytes.fromhex(env_map[action].replace(" ", ""))
        except binascii.Error:
            raise HTTPException(status_code=500, detail=f"Hex configurado invalido para acao {action}")

    # Fallback texto simples para LAN TCP/IP quando não houver protocolo binário configurado.
    if action == "message":
        return f"MESSAGE:{message}\n".encode("utf-8")
    if action == "release_entry":
        return b"RELEASE_ENTRY\n"
    if action == "release_exit":
        return b"RELEASE_EXIT\n"
    if action == "block":
        return b"BLOCK\n"
    raise HTTPException(status_code=400, detail="Acao de catraca invalida")

async def send_tcp_command(host: str, port: int, payload: bytes, timeout_seconds: float = 3.0) -> str:
    try:
        reader, writer = await asyncio.wait_for(asyncio.open_connection(host, port), timeout=timeout_seconds)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Falha ao conectar na catraca ({host}:{port}): {str(e)}")

    try:
        writer.write(payload)
        await writer.drain()
        resp = await asyncio.wait_for(reader.read(1024), timeout=timeout_seconds)
        return resp.decode("utf-8", errors="ignore")
    except asyncio.TimeoutError:
        return ""
    finally:
        writer.close()
        await writer.wait_closed()

async def get_current_user(request: Request):
    session_token = request.cookies.get("session_token")
    if session_token:
        session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
        if session:
            expires_at = session.get("expires_at")
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at and expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at and expires_at > datetime.now(timezone.utc):
                user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
                if user:
                    # verify academy billing for protected routes: owner/manager must have active trial or paid_until
                    if user.get("role") != "super_admin" and user.get("academy_id"):
                        try:
                            acad = await db.academies.find_one({"academy_id": user.get("academy_id")}, {"_id": 0})
                            if acad:
                                paid_until = acad.get("paid_until")
                                trial_until = acad.get("trial_until")
                                ok = False
                                now = datetime.now(timezone.utc)
                                if trial_until:
                                    try:
                                        t = datetime.fromisoformat(trial_until)
                                        if t.tzinfo is None:
                                            t = t.replace(tzinfo=timezone.utc)
                                        if t >= now:
                                            ok = True
                                    except Exception:
                                        pass
                                if paid_until and not ok:
                                    try:
                                        p = datetime.fromisoformat(paid_until)
                                        if p.tzinfo is None:
                                            p = p.replace(tzinfo=timezone.utc)
                                        if p >= now:
                                            ok = True
                                    except Exception:
                                        pass
                                if not ok:
                                    raise HTTPException(status_code=402, detail="Pagamento da academia necessario")
                        except HTTPException:
                            raise
                        except Exception:
                            raise HTTPException(status_code=402, detail="Pagamento da academia necessario")
                    return user

    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
        session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
        if session:
            expires_at = session.get("expires_at")
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at and expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at and expires_at > datetime.now(timezone.utc):
                user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
                if user:
                    return user
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
            user = await db.users.find_one({"user_id": payload["user_id"]}, {"_id": 0})
            if user:
                return user
        except Exception:
            pass

    raise HTTPException(status_code=401, detail="Nao autenticado")


# ============== AUTH ENDPOINTS ==============

@app.post("/api/auth/email/request-code")
async def request_email_verification(data: EmailVerificationRequest):
    email = normalize_email(data.email)
    code = f"{uuid.uuid4().int % 1000000:06d}"
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=15)
    await db.email_verifications.update_one(
        {"email": email},
        {"$set": {"email": email, "code": code, "verified": False, "expires_at": expires_at.isoformat(), "created_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    # Em produção: enviar por SMTP/provider. Aqui retornamos código para facilitar testes.
    return {"message": "Codigo enviado para validacao", "email": email, "dev_code": code}

@app.post("/api/auth/email/confirm-code")
async def confirm_email_verification(data: EmailVerificationConfirm):
    email = normalize_email(data.email)
    rec = await db.email_verifications.find_one({"email": email}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Codigo nao encontrado")
    if rec.get("code") != data.code:
        raise HTTPException(status_code=400, detail="Codigo invalido")
    try:
        exp = datetime.fromisoformat(rec.get("expires_at"))
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        if exp < datetime.now(timezone.utc):
            raise HTTPException(status_code=400, detail="Codigo expirado")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="Codigo invalido")

    await db.email_verifications.update_one({"email": email}, {"$set": {"verified": True, "verified_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Email validado", "email": email}

@app.post("/api/auth/register")
async def register(data: UserRegister, response: Response):
    email = normalize_email(data.email)
    if len((data.password or "")) < 8:
        raise HTTPException(status_code=400, detail="Senha deve ter no minimo 8 caracteres")
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    verify = await db.email_verifications.find_one({"email": email}, {"_id": 0})
    if not verify or not verify.get("verified"):
        raise HTTPException(status_code=400, detail="Email nao validado")
    if existing:
        raise HTTPException(status_code=400, detail="Email ja cadastrado")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    hashed = pwd_context.hash(data.password)
    # create user but require email verification before allowing login
    await db.users.insert_one({
        "user_id": user_id, "email": email, "name": data.name,
        "password": hashed, "role": "admin", "picture": "",
        "academy_id": "", "created_at": datetime.now(timezone.utc).isoformat(),
        "failed_login_attempts": 0, "lock_until": None,
        "email_verified": False,
    })

    # create verification token
    vtoken = binascii.hexlify(os.urandom(16)).decode()
    expires = (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
    await db.email_verifications.insert_one({
        "token": vtoken, "user_id": user_id, "email": email,
        "expires_at": expires, "created_at": datetime.now(timezone.utc).isoformat(),
    })

    verify_link = f"{BACKEND_PUBLIC_URL.rstrip('/')}/api/auth/verify-email?token={vtoken}"
    body = f"Olá {data.name},\n\nClique no link abaixo para verificar seu e-mail e ativar sua conta:\n\n{verify_link}\n\nO link expira em 24 horas.\n\nObrigado,\nGymBro"
    send_email(email, "Verificação de e-mail - GymBro", body)

    return {"message": "Registro realizado. Verifique seu e-mail para ativar a conta."}


@app.post("/api/auth/login/start")
async def login_start(data: UserLogin):
    """Step 1 of login: validate email+password, check payment and send OTP to email."""
    email = normalize_email(data.email)
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not pwd_context.verify(data.password, user.get("password", "")):
        raise HTTPException(status_code=401, detail="Credenciais invalidas")
    if not user.get("email_verified"):
        raise HTTPException(status_code=403, detail="Email nao verificado. Verifique seu e-mail antes de entrar.")

    # check academy billing
    if user.get("role") != "super_admin" and user.get("academy_id"):
        acad = await db.academies.find_one({"academy_id": user.get("academy_id")}, {"_id": 0})
        paid_until = acad.get("paid_until") if acad else None
        trial_until = acad.get("trial_until") if acad else None
        now = datetime.now(timezone.utc)
        ok = False
        if trial_until:
            try:
                t = datetime.fromisoformat(trial_until)
                if t.tzinfo is None:
                    t = t.replace(tzinfo=timezone.utc)
                if t >= now:
                    ok = True
            except Exception:
                pass
        if paid_until and not ok:
            try:
                p = datetime.fromisoformat(paid_until)
                if p.tzinfo is None:
                    p = p.replace(tzinfo=timezone.utc)
                if p >= now:
                    ok = True
            except Exception:
                pass
        if not ok:
            raise HTTPException(status_code=402, detail="Pagamento da academia necessario")

    # create challenge
    challenge_id = f"chal_{uuid.uuid4().hex[:12]}"
    code = f"{(binascii.hexlify(os.urandom(3)).hex())[:6]}"
    expires = (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat()
    await db.login_challenges.insert_one({
        "challenge_id": challenge_id, "user_id": user.get("user_id"), "email": email,
        "code": code, "expires_at": expires, "created_at": datetime.now(timezone.utc).isoformat(),
    })
    body = f"Seu codigo de acesso: {code}\n\nEste codigo expira em 10 minutos."
    send_email(email, "Codigo de acesso - GymBro", body)
    return {"challenge_id": challenge_id, "message": "Codigo enviado por e-mail"}


@app.post("/api/auth/login/verify")
async def login_verify(payload: Request, response: Response):
    body = await payload.json()
    challenge_id = body.get("challenge_id")
    code = body.get("code")
    if not challenge_id or not code:
        raise HTTPException(status_code=400, detail="challenge_id e code obrigatorios")
    doc = await db.login_challenges.find_one({"challenge_id": challenge_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Challenge invalido")
    try:
        exp = datetime.fromisoformat(doc.get("expires_at"))
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        if exp < datetime.now(timezone.utc):
            await db.login_challenges.delete_one({"challenge_id": challenge_id})
            raise HTTPException(status_code=410, detail="Challenge expirado")
    except HTTPException:
        raise
    except Exception:
        pass
    if doc.get("code") != str(code):
        raise HTTPException(status_code=401, detail="Codigo invalido")
    user = await db.users.find_one({"user_id": doc.get("user_id")}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    # issue token + session
    token = create_token(user["user_id"], user["email"])
    await db.user_sessions.insert_one({
        "user_id": user["user_id"], "session_token": token,
        "expires_at": datetime.now(timezone.utc) + timedelta(days=7),
        "created_at": datetime.now(timezone.utc),
    })
    await db.login_challenges.delete_one({"challenge_id": challenge_id})
    response.set_cookie(key="session_token", value=token, httponly=True, secure=True, samesite="none", max_age=7*24*60*60)
    return {"token": token, "user": {"user_id": user["user_id"], "email": user["email"], "name": user["name"], "role": user.get("role", "admin"), "academy_id": user.get("academy_id", "")}}

@app.post("/api/auth/login")
async def login(data: UserLogin, response: Response):
    email = normalize_email(data.email)
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if user and user.get("lock_until"):
        try:
            lock_until = user["lock_until"]
            if isinstance(lock_until, str):
                lock_until = datetime.fromisoformat(lock_until)
            if lock_until.tzinfo is None:
                lock_until = lock_until.replace(tzinfo=timezone.utc)
            if lock_until > datetime.now(timezone.utc):
                mins = int((lock_until - datetime.now(timezone.utc)).total_seconds() // 60) + 1
                raise HTTPException(status_code=423, detail=f"Conta bloqueada temporariamente. Tente novamente em {mins} minuto(s)")
        except HTTPException:
            raise
        except Exception:
            pass

    if not user or not pwd_context.verify(data.password, user.get("password", "")):
        if user:
            failed_attempts = int(user.get("failed_login_attempts", 0)) + 1
            lock_until = None
            if failed_attempts >= LOGIN_MAX_ATTEMPTS:
                lock_until = datetime.now(timezone.utc) + timedelta(minutes=LOGIN_LOCK_MINUTES)
                failed_attempts = 0
            await db.users.update_one(
                {"user_id": user["user_id"]},
                {"$set": {"failed_login_attempts": failed_attempts, "lock_until": lock_until}}
            )
        raise HTTPException(status_code=401, detail="Credenciais invalidas")

    if user.get("academy_id") and user.get("role") != "super_admin":
        has_credit = await academy_payment_is_active(user.get("academy_id", ""))
        if not has_credit:
            raise HTTPException(status_code=402, detail="Pagamento da academia pendente. Renove o credito mensal para liberar login")

    await db.users.update_one(
        {"user_id": user["user_id"]},
        {"$set": {"failed_login_attempts": 0, "lock_until": None, "last_login_at": datetime.now(timezone.utc)}}
    )

    # require email verification
    if not user.get("email_verified"):
        raise HTTPException(status_code=403, detail="Email nao verificado. Verifique seu e-mail antes de entrar.")

    # require academy payment credit for non-super admins
    if user.get("role") != "super_admin" and user.get("academy_id"):
        acad = await db.academies.find_one({"academy_id": user.get("academy_id")}, {"_id": 0})
        paid_until = None
        if acad:
            paid_until = acad.get("paid_until")
        if not paid_until:
            raise HTTPException(status_code=403, detail="Pagamento da academia nao liberado")
        try:
            paid_dt = datetime.fromisoformat(paid_until)
            if paid_dt.tzinfo is None:
                paid_dt = paid_dt.replace(tzinfo=timezone.utc)
            if paid_dt < datetime.now(timezone.utc):
                raise HTTPException(status_code=403, detail="Pagamento da academia expirado")
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=403, detail="Pagamento da academia nao liberado")
    token = create_token(user["user_id"], user["email"])
    await db.user_sessions.insert_one({
        "user_id": user["user_id"], "session_token": token,
        "expires_at": datetime.now(timezone.utc) + timedelta(days=7),
        "created_at": datetime.now(timezone.utc),
    })
    response.set_cookie(key="session_token", value=token, httponly=True, secure=True, samesite="none", max_age=7*24*60*60)
    return {"token": token, "user": {
        "user_id": user["user_id"], "email": user["email"], "name": user["name"],
        "role": user.get("role", "admin"), "academy_id": user.get("academy_id", ""),
        "picture": user.get("picture", ""),
    }}


@app.get("/api/auth/verify-email")
async def verify_email(token: str = ""):
    if not token:
        raise HTTPException(status_code=400, detail="Token obrigatorio")
    doc = await db.email_verifications.find_one({"token": token}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Token invalido ou expirado")
    try:
        exp = datetime.fromisoformat(doc.get("expires_at"))
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        if exp < datetime.now(timezone.utc):
            await db.email_verifications.delete_one({"token": token})
            raise HTTPException(status_code=410, detail="Token expirado")
    except HTTPException:
        raise
    except Exception:
        pass

    await db.users.update_one({"user_id": doc.get("user_id")}, {"$set": {"email_verified": True}})
    await db.email_verifications.delete_one({"token": token})
    return {"message": "Email verificado com sucesso"}

@app.post("/api/payments/academy/subscription/checkout")
async def create_academy_subscription_checkout(data: AcademySubscriptionCheckout, user=Depends(get_current_user)):
    academy = await db.academies.find_one({"academy_id": data.academy_id}, {"_id": 0})
    if not academy:
        raise HTTPException(status_code=404, detail="Academia nao encontrada")
    if data.amount <= 0:
        raise HTTPException(status_code=400, detail="Valor invalido")
    if not MP_ACCESS_TOKEN or MP_ACCESS_TOKEN == "placeholder":
        raise HTTPException(status_code=503, detail="Mercado Pago nao configurado")

    cycle_ref = datetime.now(timezone.utc).strftime("%Y-%m")
    external_reference = f"academy:{data.academy_id}:{cycle_ref}"
    preference_payload = {
        "items": [{
            "title": data.description or f"Mensalidade {academy.get('nome', 'Academia')}",
            "quantity": 1,
            "currency_id": "BRL",
            "unit_price": round(float(data.amount), 2),
        }],
        "payer": {"email": data.payer_email or user.get("email", "")},
        "external_reference": external_reference,
        "back_urls": {
            "success": f"{FRONTEND_URL}/admin?payment=success",
            "failure": f"{FRONTEND_URL}/admin?payment=failure",
            "pending": f"{FRONTEND_URL}/admin?payment=pending",
        },
        "auto_return": "approved",
        "notification_url": f"{BACKEND_PUBLIC_URL.rstrip('/')}/api/webhooks/mercadopago",
    }

    async with httpx.AsyncClient(timeout=20) as http_client:
        resp = await http_client.post(
            "https://api.mercadopago.com/checkout/preferences",
            headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}", "Content-Type": "application/json"},
            json=preference_payload,
        )
    if resp.status_code not in (200, 201):
        raise HTTPException(status_code=502, detail=f"Erro ao criar checkout MP: {resp.text[:200]}")
    pref = resp.json()

    await db.academy_billing.insert_one({
        "billing_id": f"bill_{uuid.uuid4().hex[:12]}",
        "academy_id": data.academy_id,
        "academy_name": academy.get("nome", ""),
        "external_reference": external_reference,
        "preference_id": pref.get("id", ""),
        "status": "pending",
        "amount": round(float(data.amount), 2),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("user_id", ""),
    })

    return {
        "checkout_url": pref.get("init_point", ""),
        "sandbox_checkout_url": pref.get("sandbox_init_point", ""),
        "external_reference": external_reference,
        "preference_id": pref.get("id", ""),
    }

@app.post("/api/auth/google/session")
async def google_session(request: Request, response: Response):
    body = await request.json()
    session_id = body.get("session_id")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id obrigatorio")
    # REMINDER: DO NOT HARDCODE THE URL, OR ADD ANY FALLBACKS OR REDIRECT URLS, THIS BREAKS THE AUTH
    async with httpx.AsyncClient() as client_http:
        resp = await client_http.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": session_id},
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=401, detail="Sessao invalida")
        data = resp.json()

    email = data["email"]
    name = data.get("name", "")
    picture = data.get("picture", "")
    session_token = data.get("session_token", f"st_{uuid.uuid4().hex}")

    existing = await db.users.find_one({"email": email}, {"_id": 0})
    verify = await db.email_verifications.find_one({"email": email}, {"_id": 0})
    if not verify or not verify.get("verified"):
        raise HTTPException(status_code=400, detail="Email nao validado")
    if existing:
        user_id = existing["user_id"]
        await db.users.update_one({"user_id": user_id}, {"$set": {"name": name, "picture": picture}})
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": user_id, "email": email, "name": name, "picture": picture,
            "password": "", "role": "admin", "academy_id": "", "created_at": datetime.now(timezone.utc),
        })

    await db.user_sessions.insert_one({
        "user_id": user_id, "session_token": session_token,
        "expires_at": datetime.now(timezone.utc) + timedelta(days=7),
        "created_at": datetime.now(timezone.utc),
    })

    current_user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if current_user and current_user.get("academy_id") and current_user.get("role") != "super_admin":
        has_credit = await academy_payment_is_active(current_user.get("academy_id", ""))
        if not has_credit:
            raise HTTPException(status_code=402, detail="Pagamento da academia pendente. Renove o credito mensal para liberar login")

    response.set_cookie(key="session_token", value=session_token, httponly=True,
                        secure=True, samesite="none", path="/", max_age=7*24*3600)
    return {
        "user": {"user_id": user_id, "email": email, "name": name, "picture": picture, "role": current_user.get("role", "admin") if current_user else "admin", "academy_id": current_user.get("academy_id", "") if current_user else ""},
        "session_token": session_token,
    }

@app.get("/api/auth/me")
async def auth_me(user=Depends(get_current_user)):
    if user.get("academy_id") and user.get("role") != "super_admin":
        has_credit = await academy_payment_is_active(user.get("academy_id", ""))
        if not has_credit:
            raise HTTPException(status_code=402, detail="Pagamento da academia pendente")
    return {
        "user_id": user["user_id"], "email": user["email"], "name": user["name"],
        "picture": user.get("picture", ""), "role": user.get("role", "admin"),
        "academy_id": user.get("academy_id", ""),
    }

@app.post("/api/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_many({"session_token": session_token})
    response.delete_cookie("session_token", path="/")
    return {"message": "Logout realizado"}

# ============== STUDENTS CRUD ==============

@app.get("/api/students")
async def list_students(user=Depends(get_current_user), search: str = "", status: str = "", academy_id: str = ""):
    query = {}
    if search:
        query["$or"] = [
            {"nome": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"cpf": {"$regex": search, "$options": "i"}},
            {"tag_rfid": {"$regex": search, "$options": "i"}},
        ]
    if status:
        query["status"] = status
    if academy_id:
        query["academy_id"] = academy_id
    elif user.get("role") != "super_admin" and user.get("academy_id"):
        query["academy_id"] = user["academy_id"]
    students = await db.students.find(query, {"_id": 0}).sort("nome", 1).to_list(1000)
    return students

@app.post("/api/students")
async def create_student(data: StudentCreate, user=Depends(get_current_user)):
    existing = await db.students.find_one({"cpf": data.cpf}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="CPF ja cadastrado")
    student_id = f"std_{uuid.uuid4().hex[:12]}"
    doc = {"student_id": student_id, **data.model_dump(),
           "created_at": datetime.now(timezone.utc).isoformat(),
           "updated_at": datetime.now(timezone.utc).isoformat()}
    if not doc.get("academy_id") and user.get("academy_id"):
        doc["academy_id"] = user["academy_id"]
    await db.students.insert_one(doc)
    return await db.students.find_one({"student_id": student_id}, {"_id": 0})

@app.get("/api/students/{student_id}")
async def get_student(student_id: str, user=Depends(get_current_user)):
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    return student

@app.put("/api/students/{student_id}")
async def update_student(student_id: str, data: StudentUpdate, user=Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum dado para atualizar")
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.students.update_one({"student_id": student_id}, {"$set": update_data})
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    return student

@app.post("/api/students/{student_id}/biometria")
async def register_student_biometria(student_id: str, data: BiometricRegister, user=Depends(get_current_user)):
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    existing = await db.students.find_one({"biometria_id": data.biometria_id, "student_id": {"$ne": student_id}}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Biometria ja cadastrada para outro aluno")
    await db.students.update_one({"student_id": student_id}, {"$set": {"biometria_id": data.biometria_id, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return await db.students.find_one({"student_id": student_id}, {"_id": 0})

@app.delete("/api/students/{student_id}")
async def delete_student(student_id: str, user=Depends(get_current_user)):
    result = await db.students.delete_one({"student_id": student_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    return {"message": "Aluno removido"}

# ============== PLANS CRUD ==============

@app.get("/api/plans")
async def list_plans(user=Depends(get_current_user)):
    return await db.plans.find({}, {"_id": 0}).sort("valor", 1).to_list(100)

@app.get("/api/plans/public")
async def list_plans_public():
    return await db.plans.find({"ativo": True}, {"_id": 0}).sort("valor", 1).to_list(100)

@app.post("/api/plans")
async def create_plan(data: PlanCreate, user=Depends(get_current_user)):
    plan_id = f"plan_{uuid.uuid4().hex[:12]}"
    doc = {"plan_id": plan_id, **data.model_dump(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.plans.insert_one(doc)
    return await db.plans.find_one({"plan_id": plan_id}, {"_id": 0})

@app.put("/api/plans/{plan_id}")
async def update_plan(plan_id: str, data: PlanUpdate, user=Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum dado para atualizar")
    await db.plans.update_one({"plan_id": plan_id}, {"$set": update_data})
    plan = await db.plans.find_one({"plan_id": plan_id}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plano nao encontrado")
    return plan

@app.delete("/api/plans/{plan_id}")
async def delete_plan(plan_id: str, user=Depends(get_current_user)):
    result = await db.plans.delete_one({"plan_id": plan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plano nao encontrado")
    return {"message": "Plano removido"}

# ============== DASHBOARD ==============

@app.get("/api/dashboard")
async def dashboard_stats(user=Depends(get_current_user)):
    q = {}
    if user.get("role") != "super_admin" and user.get("academy_id"):
        q["academy_id"] = user["academy_id"]

    total_students = await db.students.count_documents(q)
    active_q = {**q, "status": "ativo"}
    inactive_q = {**q, "status": "inativo"}
    active_students = await db.students.count_documents(active_q)
    inactive_students = await db.students.count_documents(inactive_q)
    total_plans = await db.plans.count_documents({})

    pipeline = [{"$match": active_q},
                {"$lookup": {"from": "plans", "localField": "plano_id", "foreignField": "plan_id", "as": "plano"}},
                {"$unwind": {"path": "$plano", "preserveNullAndEmptyArrays": True}},
                {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$plano.valor", 0]}}}}]
    rev = await db.students.aggregate(pipeline).to_list(1)
    faturamento = rev[0]["total"] if rev else 0

    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    log_q = {"timestamp": {"$gte": today_start.isoformat()}}
    if q.get("academy_id"):
        log_q["academy_id"] = q["academy_id"]
    acessos_hoje = await db.access_logs.count_documents(log_q)

    recent_q = {}
    if q.get("academy_id"):
        recent_q["academy_id"] = q["academy_id"]
    recent_logs = await db.access_logs.find(recent_q, {"_id": 0}).sort("timestamp", -1).limit(10).to_list(10)

    # Occupancy estimate (entries in last hour)
    hour_ago = (datetime.now(timezone.utc) - timedelta(hours=1)).isoformat()
    occ_q = {"timestamp": {"$gte": hour_ago}, "autorizado": True}
    if q.get("academy_id"):
        occ_q["academy_id"] = q["academy_id"]
    ocupacao_atual = await db.access_logs.count_documents(occ_q)

    return {
        "total_alunos": total_students, "alunos_ativos": active_students,
        "alunos_inativos": inactive_students, "total_planos": total_plans,
        "faturamento_mensal": faturamento, "acessos_hoje": acessos_hoje,
        "ocupacao_atual": ocupacao_atual, "ultimos_acessos": recent_logs,
    }

# ============== CHART DATA ==============

@app.get("/api/dashboard/charts")
async def dashboard_charts(user=Depends(get_current_user)):
    q = {}
    if user.get("role") != "super_admin" and user.get("academy_id"):
        q["academy_id"] = user["academy_id"]

    # Revenue by plan
    pipeline = [{"$match": {**q, "status": "ativo"}},
                {"$lookup": {"from": "plans", "localField": "plano_id", "foreignField": "plan_id", "as": "plano"}},
                {"$unwind": {"path": "$plano", "preserveNullAndEmptyArrays": True}},
                {"$group": {"_id": "$plano.nome", "total": {"$sum": {"$ifNull": ["$plano.valor", 0]}}, "count": {"$sum": 1}}}]
    rev_by_plan = await db.students.aggregate(pipeline).to_list(20)
    revenue_chart = [{"plano": r["_id"] or "Sem Plano", "valor": r["total"], "alunos": r["count"]} for r in rev_by_plan]

    # Access by hour (last 24h)
    now = datetime.now(timezone.utc)
    hours_data = []
    for i in range(24):
        h = now - timedelta(hours=23-i)
        h_start = h.replace(minute=0, second=0, microsecond=0)
        h_end = h_start + timedelta(hours=1)
        hq = {"timestamp": {"$gte": h_start.isoformat(), "$lt": h_end.isoformat()}}
        if q.get("academy_id"):
            hq["academy_id"] = q["academy_id"]
        cnt = await db.access_logs.count_documents(hq)
        hours_data.append({"hora": h_start.strftime("%H:00"), "acessos": cnt})

    # Students by status
    status_chart = [
        {"status": "Ativos", "count": await db.students.count_documents({**q, "status": "ativo"})},
        {"status": "Inativos", "count": await db.students.count_documents({**q, "status": "inativo"})},
    ]

    # Monthly revenue (last 6 months simulated from active students)
    monthly = []
    for i in range(6):
        month_date = now - timedelta(days=30*i)
        monthly.append({
            "mes": month_date.strftime("%b/%y"),
            "valor": round(revenue_chart and sum(r["valor"] for r in revenue_chart) * (0.7 + 0.05*i) or 0, 2),
        })
    monthly.reverse()

    return {
        "receita_por_plano": revenue_chart,
        "acessos_por_hora": hours_data,
        "alunos_por_status": status_chart,
        "receita_mensal": monthly,
    }

# ============== ACCESS LOGS ==============

@app.get("/api/access-logs")
async def list_access_logs(user=Depends(get_current_user), limit: int = 50):
    q = {}
    if user.get("role") != "super_admin" and user.get("academy_id"):
        q["academy_id"] = user["academy_id"]
    return await db.access_logs.find(q, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)

# ============== ACCESS VALIDATION (for Local Agent) ==============

@app.post("/api/access/validate")
async def validate_access(data: AccessValidation):
    query = {}
    if data.tipo == "rfid":
        query["tag_rfid"] = data.tag_id
    elif data.tipo == "biometria":
        query["biometria_id"] = data.tag_id
    elif data.tipo == "teclado":
        query["cpf"] = data.tag_id
    else:
        query["tag_rfid"] = data.tag_id

    student = await db.students.find_one(query, {"_id": 0})

    log_base = {
        "log_id": f"log_{uuid.uuid4().hex[:12]}",
        "tag_id": data.tag_id, "tipo": data.tipo,
        "academy_id": data.academy_id or (student or {}).get("academy_id", ""),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }

    if not student:
        log_base.update({"student_id": "", "student_name": "Desconhecido", "autorizado": False, "motivo": "Aluno nao encontrado"})
        await db.access_logs.insert_one(log_base)
        await ws_manager.broadcast({"type": "access", "data": {k: v for k, v in log_base.items() if k != "_id"}})
        return {"autorizado": False, "motivo": "Aluno nao encontrado", "aluno": None}

    if student.get("status") != "ativo":
        log_base.update({"student_id": student["student_id"], "student_name": student["nome"], "autorizado": False, "motivo": "Assinatura inativa"})
        await db.access_logs.insert_one(log_base)
        await ws_manager.broadcast({"type": "access", "data": {k: v for k, v in log_base.items() if k != "_id"}})
        return {"autorizado": False, "motivo": "Assinatura inativa", "aluno": student["nome"]}

    vencimento = student.get("data_vencimento", "")
    if vencimento:
        try:
            dt_venc = datetime.fromisoformat(vencimento)
            if dt_venc.tzinfo is None:
                dt_venc = dt_venc.replace(tzinfo=timezone.utc)
            if dt_venc < datetime.now(timezone.utc):
                await db.students.update_one({"student_id": student["student_id"]}, {"$set": {"status": "inativo"}})
                log_base.update({"student_id": student["student_id"], "student_name": student["nome"], "autorizado": False, "motivo": "Assinatura vencida"})
                await db.access_logs.insert_one(log_base)
                await ws_manager.broadcast({"type": "access", "data": {k: v for k, v in log_base.items() if k != "_id"}})
                return {"autorizado": False, "motivo": "Assinatura vencida", "aluno": student["nome"]}
        except Exception:
            pass

    log_base.update({"student_id": student["student_id"], "student_name": student["nome"], "autorizado": True, "motivo": "Acesso liberado"})
    await db.access_logs.insert_one(log_base)
    await ws_manager.broadcast({"type": "access", "data": {k: v for k, v in log_base.items() if k != "_id"}})
    return {"autorizado": True, "motivo": "Acesso liberado", "aluno": student["nome"]}

# ============== MERCADO PAGO WEBHOOK ==============

@app.post("/api/webhooks/mercadopago")
async def mercadopago_webhook(request: Request):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Payload invalido")

    action = body.get("action", "")
    data = body.get("data", {})
    payment_id = data.get("id")

    if action not in ("payment.created", "payment.updated"):
        return {"status": "ignored", "reason": "Acao nao relevante"}
    if not payment_id:
        return {"status": "ignored", "reason": "Sem ID de pagamento"}


    if not MP_ACCESS_TOKEN or MP_ACCESS_TOKEN == "placeholder":
        await db.webhook_logs.insert_one({
            "log_id": f"wh_{uuid.uuid4().hex[:12]}", "action": action,
            "payment_id": str(payment_id), "status": "token_not_configured",
            "timestamp": datetime.now(timezone.utc).isoformat(),
        })
        return {"status": "received", "message": "Token MP nao configurado - webhook registrado"}

    try:
        async with httpx.AsyncClient() as http_client:
            resp = await http_client.get(f"https://api.mercadopago.com/v1/payments/{payment_id}",
                                         headers={"Authorization": f"Bearer {MP_ACCESS_TOKEN}"})
            if resp.status_code != 200:
                await db.webhook_logs.insert_one({
                    "log_id": f"wh_{uuid.uuid4().hex[:12]}", "action": action,
                    "payment_id": str(payment_id), "status": "mp_api_error", "error": resp.text,
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                })
                return {"status": "error", "message": "Erro ao consultar MP"}
            payment_data = resp.json()
    except Exception as e:
        await db.webhook_logs.insert_one({
            "log_id": f"wh_{uuid.uuid4().hex[:12]}", "action": action,
            "payment_id": str(payment_id), "status": "exception", "error": str(e),
            "timestamp": datetime.now(timezone.utc).isoformat(),
        })
        return {"status": "error", "message": str(e)}

    payment_status = payment_data.get("status")
    payer_email = payment_data.get("payer", {}).get("email", "")
    external_ref = payment_data.get("external_reference", "")

    await db.webhook_logs.insert_one({
        "log_id": f"wh_{uuid.uuid4().hex[:12]}", "action": action,
        "payment_id": str(payment_id), "payment_status": payment_status,
        "payer_email": payer_email, "external_reference": external_ref,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    })

    if payment_status != "approved":
        return {"status": "received", "payment_status": payment_status}

    if external_ref.startswith("academy:"):
        parts = external_ref.split(":")
        academy_id = parts[1] if len(parts) > 1 else ""
        if academy_id:
            academy = await db.academies.find_one({"academy_id": academy_id}, {"_id": 0})
            if academy:
                payment_dt = datetime.now(timezone.utc)
                doc = await db.academy_billing.find_one({"external_reference": external_ref}, {"_id": 0})
                paid_until = payment_dt + timedelta(days=30)
                if doc and doc.get("paid_until"):
                    try:
                        prev = datetime.fromisoformat(doc["paid_until"])
                        if prev.tzinfo is None:
                            prev = prev.replace(tzinfo=timezone.utc)
                        base = prev if prev > payment_dt else payment_dt
                        paid_until = base + timedelta(days=30)
                    except Exception:
                        pass
                await db.academy_billing.update_one(
                    {"external_reference": external_ref},
                    {"$set": {
                        "academy_id": academy_id,
                        "academy_name": academy.get("nome", ""),
                        "status": payment_status,
                        "payment_id": str(payment_id),
                        "payment_status": payment_status,
                        "payer_email": payer_email,
                        "paid_until": paid_until.isoformat(),
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }},
                    upsert=True,
                )
                # update academy paid_until and record payment
                try:
                    await db.academies.update_one({"academy_id": academy_id}, {"$set": {"paid_until": paid_until.isoformat()}})
                    await db.payments.insert_one({
                        "payment_id": str(payment_id), "academy_id": academy_id,
                        "external_reference": external_ref, "amount": payment_data.get("transaction_amount", 0),
                        "status": payment_status, "payer_email": payer_email,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    })
                except Exception:
                    pass
                return {"status": "processed", "message": f"Mensalidade da academia paga ate {paid_until.strftime('%d/%m/%Y')}"}

    student = None
    if external_ref:
        student = await db.students.find_one({"student_id": external_ref}, {"_id": 0})
    if not student and payer_email:
        student = await db.students.find_one({"email": payer_email}, {"_id": 0})
    if not student:
        return {"status": "received", "message": "Pagamento aprovado mas aluno nao encontrado"}

    plan = None
    if student.get("plano_id"):
        plan = await db.plans.find_one({"plan_id": student["plano_id"]}, {"_id": 0})
    days_to_add = plan["duracao_dias"] if plan else 30

    current_expiry = student.get("data_vencimento", "")
    try:
        base_date = datetime.fromisoformat(current_expiry) if current_expiry else datetime.now(timezone.utc)
        if base_date.tzinfo is None:
            base_date = base_date.replace(tzinfo=timezone.utc)
        if base_date < datetime.now(timezone.utc):
            base_date = datetime.now(timezone.utc)
    except Exception:
        base_date = datetime.now(timezone.utc)

    new_expiry = base_date + timedelta(days=days_to_add)
    await db.students.update_one(
        {"student_id": student["student_id"]},
        {"$set": {"status": "ativo", "data_vencimento": new_expiry.isoformat(),
                  "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"status": "processed", "message": f"Assinatura renovada ate {new_expiry.strftime('%d/%m/%Y')}", "aluno": student["nome"]}

@app.get("/api/webhook-logs")
async def list_webhook_logs(user=Depends(get_current_user), limit: int = 50):
    return await db.webhook_logs.find({}, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)

# ============== ACADEMIES (MULTI-TENANCY) ==============

@app.get("/api/academies")
async def list_academies(user=Depends(get_current_user)):
    return await db.academies.find({}, {"_id": 0}).sort("nome", 1).to_list(100)

@app.post("/api/academies")
@app.post("/api/academies")
async def create_academy(data: AcademyCreate, user=Depends(get_current_user)):
<<<<<<< HEAD
    count = await db.academies.count_documents({})
    if count >= 1:
        raise HTTPException(status_code=400, detail="Apenas uma filial/academia permitida nesta versao")
=======
    # enforce single academy per owner (não é multi-tenancy, é single-tenant por dono)
    if user.get("role") != "super_admin":
        # Check if user already linked to an academy
        if user.get("academy_id"):
            existing = await db.academies.find_one({"academy_id": user.get("academy_id")}, {"_id": 0})
            if existing:
                raise HTTPException(status_code=400, detail="Voce ja possui uma filial. Cada usuario pode ter apenas uma academia registrada no sistema.")
        
        # Double-check: also check academies.owner_user_id
        existing = await db.academies.find_one({"owner_user_id": user.get("user_id")}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail="Voce ja possui uma filial. Cada usuario pode ter apenas uma academia registrada no sistema.")

>>>>>>> d3764ba4 (versão 2.0)
    academy_id = f"acad_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    trial_until = (now + timedelta(days=30)).isoformat()
    doc = {"academy_id": academy_id, **data.model_dump(),
           "created_at": now.isoformat(), "owner_user_id": user.get("user_id"),
           "paid_until": None, "trial_until": trial_until, "billing_status": "trial"}
    await db.academies.insert_one(doc)
    # link user to academy (owner)
    try:
        await db.users.update_one({"user_id": user.get("user_id")}, {"$set": {"academy_id": academy_id}})
    except Exception:
        pass
    return await db.academies.find_one({"academy_id": academy_id}, {"_id": 0})

@app.put("/api/academies/{academy_id}")
async def update_academy(academy_id: str, data: AcademyUpdate, user=Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum dado para atualizar")
    await db.academies.update_one({"academy_id": academy_id}, {"$set": update_data})
    acad = await db.academies.find_one({"academy_id": academy_id}, {"_id": 0})
    if not acad:
        raise HTTPException(status_code=404, detail="Academia nao encontrada")
    return acad

@app.delete("/api/academies/{academy_id}")
async def delete_academy(academy_id: str, user=Depends(get_current_user)):
    result = await db.academies.delete_one({"academy_id": academy_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Academia nao encontrada")
    return {"message": "Academia removida"}


@app.post("/api/students/{student_id}/passkey/register")
async def register_student_passkey(student_id: str, request: Request, user=Depends(get_current_user)):
    """Store a student's WebAuthn/public-key credential (publicKey) as a passkey.
    Body: { credential: { id, publicKey, transports?, type } }
    """
    body = await request.json()
    cred = body.get("credential")
    if not cred:
        raise HTTPException(status_code=400, detail="credential obrigatorio")
    stud = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not stud:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    creds = stud.get("webauthn_credentials", []) or []
    creds.append({"id": cred.get("id"), "public_key": cred.get("publicKey"), "created_at": datetime.now(timezone.utc).isoformat()})
    await db.students.update_one({"student_id": student_id}, {"$set": {"webauthn_credentials": creds}})
    return {"message": "Passkey registrada"}


@app.post("/api/students/{student_id}/passkey/register/options")
async def passkey_register_options(student_id: str, user=Depends(get_current_user)):
    """Begin WebAuthn registration: returns PublicKeyCredentialCreationOptions and stores state."""
    stud = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not stud:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    user_obj = {
        "id": stud["student_id"].encode('utf-8'),
        "name": stud.get("email") or stud.get("nome"),
        "displayName": stud.get("nome") or stud.get("email"),
    }
    credentials = []
    # existing credentials to prevent duplicates
    existing = stud.get("webauthn_credentials") or []
    for c in existing:
        try:
            credentials.append({"type": "public-key", "id": c.get("id")})
        except Exception:
            pass
    reg_data, state = fido_server.register_begin(user_obj, credentials=credentials)
    # store state as base64 cbor
    try:
        state_bytes = cbor.dumps(state)
        state_b64 = binascii.b2a_base64(state_bytes).decode().strip()
        state_id = f"wr_{uuid.uuid4().hex[:12]}"
        await db.webauthn_registrations.insert_one({"state_id": state_id, "student_id": student_id, "state": state_b64, "created_at": datetime.now(timezone.utc).isoformat()})
    except Exception:
        raise HTTPException(status_code=500, detail="Falha ao criar desafio WebAuthn")
    # registration options (reg_data) is JSON-serializable but bytes need encoding
    # fido2 returns bytes for challenge and user.id; convert to base64 URL-safe
    def encode_buf(b):
        return binascii.b2a_base64(b).decode().strip()
    if isinstance(reg_data.get('challenge'), (bytes, bytearray)):
        reg_data['challenge'] = encode_buf(reg_data['challenge'])
    if reg_data.get('user') and isinstance(reg_data['user'].get('id'), (bytes, bytearray)):
        reg_data['user']['id'] = encode_buf(reg_data['user']['id'])
    return {"state_id": state_id, "publicKey": reg_data}


@app.post("/api/students/{student_id}/passkey/register/verify")
async def passkey_register_verify(student_id: str, request: Request, user=Depends(get_current_user)):
    body = await request.json()
    state_id = body.get('state_id')
    rawId_b64 = body.get('rawId')
    clientDataJSON_b64 = body.get('clientDataJSON')
    attestation_b64 = body.get('publicKey') or body.get('attestationObject')
    if not state_id or not rawId_b64 or not clientDataJSON_b64 or not attestation_b64:
        raise HTTPException(status_code=400, detail='Parametros obrigatorios: state_id, rawId, clientDataJSON, publicKey')
    reg = await db.webauthn_registrations.find_one({'state_id': state_id}, {'_id': 0})
    if not reg:
        raise HTTPException(status_code=404, detail='State nao encontrado')
    try:
        state_bytes = binascii.a2b_base64(reg['state'])
        state = cbor.loads(state_bytes)
    except Exception:
        raise HTTPException(status_code=500, detail='State invalido')
    try:
        rawId = binascii.a2b_base64(rawId_b64)
        clientDataJSON = binascii.a2b_base64(clientDataJSON_b64)
        attestationObject = binascii.a2b_base64(attestation_b64)
        client_data = CollectedClientData(clientDataJSON)
        att_obj = AttestationObject(attestationObject)
        auth_data = fido_server.register_complete(state, client_data, att_obj)
        cred_data = auth_data.credential_data
        cred_id = binascii.b2a_base64(cred_data.credential_id).decode().strip()
        # store public key as COSE/DER blob if possible
        try:
            pubkey = cred_data.public_key
            # encode COSE key to CBOR bytes
            pub_bytes = pubkey.encode()
            pub_b64 = binascii.b2a_base64(pub_bytes).decode().strip()
        except Exception:
            pub_b64 = ''
        entry = {"id": cred_id, "public_key": pub_b64, "created_at": datetime.now(timezone.utc).isoformat(), "sign_count": getattr(cred_data, 'sign_count', 0)}
        await db.students.update_one({'student_id': student_id}, {'$push': {'webauthn_credentials': entry}})
        # remove state
        await db.webauthn_registrations.delete_one({'state_id': state_id})
        return {"message": "Passkey verificada e registrada", "credential_id": cred_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Falha ao validar attestation: {str(e)}")


@app.get("/api/students/{student_id}/passkeys")
async def list_student_passkeys(student_id: str, user=Depends(get_current_user)):
    stud = await db.students.find_one({"student_id": student_id}, {"_id": 0, "webauthn_credentials": 1})
    if not stud:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    return {"webauthn_credentials": stud.get("webauthn_credentials", [])}


@app.post("/api/students/{student_id}/passkey/auth/options")
async def passkey_auth_options(student_id: str, user=Depends(get_current_user)):
    """Begin WebAuthn authentication for a student: returns options and stores state."""
    stud = await db.students.find_one({"student_id": student_id}, {"_id": 0, "webauthn_credentials": 1})
    if not stud:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    existing = stud.get("webauthn_credentials") or []
    allowed = []
    for c in existing:
        try:
            # decode stored base64 credential id to raw bytes
            cid = binascii.a2b_base64(c.get("id")) if c.get("id") else None
            if cid:
                allowed.append({"type": "public-key", "id": cid})
        except Exception:
            continue

    auth_data, state = fido_server.authenticate_begin(allowed)
    try:
        state_bytes = cbor.dumps(state)
        state_b64 = binascii.b2a_base64(state_bytes).decode().strip()
        state_id = f"wa_{uuid.uuid4().hex[:12]}"
        await db.webauthn_registrations.insert_one({"state_id": state_id, "student_id": student_id, "state": state_b64, "created_at": datetime.now(timezone.utc).isoformat()})
    except Exception:
        raise HTTPException(status_code=500, detail="Falha ao criar desafio WebAuthn")

    # encode challenge and any byte fields
    def encode_buf(b):
        return binascii.b2a_base64(b).decode().strip()
    if isinstance(auth_data.get('challenge'), (bytes, bytearray)):
        auth_data['challenge'] = encode_buf(auth_data['challenge'])
    if auth_data.get('allowCredentials'):
        for ac in auth_data['allowCredentials']:
            if isinstance(ac.get('id'), (bytes, bytearray)):
                ac['id'] = encode_buf(ac['id'])

    return {"state_id": state_id, "publicKey": auth_data}


@app.post("/api/students/{student_id}/passkey/auth/verify")
async def passkey_auth_verify(student_id: str, request: Request):
    """Verify an assertion from the client. Body must contain: state_id, rawId, clientDataJSON, authenticatorData, signature"""
    body = await request.json()
    state_id = body.get('state_id')
    rawId_b64 = body.get('rawId')
    clientDataJSON_b64 = body.get('clientDataJSON')
    authenticatorData_b64 = body.get('authenticatorData')
    signature_b64 = body.get('signature')
    if not state_id or not rawId_b64 or not clientDataJSON_b64 or not authenticatorData_b64 or not signature_b64:
        raise HTTPException(status_code=400, detail='Parametros obrigatorios: state_id, rawId, clientDataJSON, authenticatorData, signature')

    reg = await db.webauthn_registrations.find_one({'state_id': state_id}, {'_id': 0})
    if not reg:
        raise HTTPException(status_code=404, detail='State nao encontrado')
    try:
        state_bytes = binascii.a2b_base64(reg['state'])
        state = cbor.loads(state_bytes)
    except Exception:
        raise HTTPException(status_code=500, detail='State invalido')

    try:
        rawId = binascii.a2b_base64(rawId_b64)
        clientDataJSON = binascii.a2b_base64(clientDataJSON_b64)
        authenticatorData = binascii.a2b_base64(authenticatorData_b64)
        signature = binascii.a2b_base64(signature_b64)

        client_data = CollectedClientData(clientDataJSON)
        auth_data = AuthenticatorData(authenticatorData)

        res = fido_server.authenticate_complete(state, rawId, client_data, auth_data, signature)
        # res may contain credential and sign_count
        cred = getattr(res, 'credential', None)
        sign_count = getattr(res, 'signature_count', None) or getattr(res, 'sign_count', None)

        # increment stored sign_count for that credential
        if cred:
            cred_id_b64 = binascii.b2a_base64(cred.credential_id).decode().strip()
            await db.students.update_one({'student_id': student_id, 'webauthn_credentials.id': cred_id_b64}, {'$set': {'webauthn_credentials.$.sign_count': sign_count or 0}})
        # remove state
        await db.webauthn_registrations.delete_one({'state_id': state_id})
        return {'authenticated': True, 'student_id': student_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'Falha ao validar assertion: {str(e)}')

@app.get("/api/academies/{academy_id}/stats")
async def academy_stats(academy_id: str, user=Depends(get_current_user)):
    q = {"academy_id": academy_id}
    total = await db.students.count_documents(q)
    active = await db.students.count_documents({**q, "status": "ativo"})
    inactive = await db.students.count_documents({**q, "status": "inativo"})
    pipeline = [{"$match": {**q, "status": "ativo"}},
                {"$lookup": {"from": "plans", "localField": "plano_id", "foreignField": "plan_id", "as": "plano"}},
                {"$unwind": {"path": "$plano", "preserveNullAndEmptyArrays": True}},
                {"$group": {"_id": None, "total": {"$sum": {"$ifNull": ["$plano.valor", 0]}}}}]
    rev = await db.students.aggregate(pipeline).to_list(1)
    return {"total_alunos": total, "alunos_ativos": active, "alunos_inativos": inactive,
            "faturamento": rev[0]["total"] if rev else 0}

# ============== NOTIFICATIONS (MOCKED EMAIL) ==============

@app.get("/api/notifications")
async def list_notifications(user=Depends(get_current_user), limit: int = 50):
    return await db.notifications.find({}, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)

@app.post("/api/notifications/check-expiring")
async def check_expiring_subscriptions(user=Depends(get_current_user)):
    """Check for students whose subscriptions expire in the next 7 days and create notifications"""
    now = datetime.now(timezone.utc)
    week_from_now = (now + timedelta(days=7)).isoformat()
    expiring = await db.students.find({
        "status": "ativo",
        "data_vencimento": {"$lte": week_from_now, "$gte": now.isoformat()},
    }, {"_id": 0}).to_list(1000)

    created = 0
    for student in expiring:
        existing = await db.notifications.find_one({
            "student_id": student["student_id"], "tipo": "vencimento",
            "created_at": {"$gte": now.replace(hour=0, minute=0).isoformat()}
        })
        if existing:
            continue
        dt_venc = student.get("data_vencimento", "")
        try:
            venc_fmt = datetime.fromisoformat(dt_venc).strftime("%d/%m/%Y")
        except Exception:
            venc_fmt = dt_venc
        days_left = 0
        try:
            dt = datetime.fromisoformat(dt_venc)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            days_left = (dt - now).days
        except Exception:
            pass

        notif = {
            "notif_id": f"notif_{uuid.uuid4().hex[:12]}",
            "tipo": "vencimento",
            "titulo": f"Assinatura vencendo em {days_left} dias",
            "mensagem": f"O aluno {student['nome']} ({student['email']}) tem assinatura vencendo em {venc_fmt}.",
            "student_id": student["student_id"],
            "student_name": student["nome"],
            "student_email": student["email"],
            "email_enviado": True,
            "email_destino": student["email"],
            "email_status": "enviado (simulado)",
            "lida": False,
            "created_at": now.isoformat(),
        }
        await db.notifications.insert_one(notif)
        created += 1

    return {"message": f"{created} notificacoes criadas", "total_vencendo": len(expiring)}

@app.put("/api/notifications/{notif_id}/read")
async def mark_notification_read(notif_id: str, user=Depends(get_current_user)):
    await db.notifications.update_one({"notif_id": notif_id}, {"$set": {"lida": True}})
    return {"message": "Notificacao marcada como lida"}

@app.delete("/api/notifications/{notif_id}")
async def delete_notification(notif_id: str, user=Depends(get_current_user)):
    await db.notifications.delete_one({"notif_id": notif_id})
    return {"message": "Notificacao removida"}

# ============== REPORTS (PDF / EXCEL) ==============

@app.get("/api/reports/students/excel")
async def export_students_excel(user=Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    students = await db.students.find({}, {"_id": 0}).sort("nome", 1).to_list(5000)
    plans_list = await db.plans.find({}, {"_id": 0}).to_list(100)
    plans_map = {p["plan_id"]: p["nome"] for p in plans_list}

    wb = Workbook()
    ws = wb.active
    ws.title = "Alunos"
    headers = ["Nome", "Email", "CPF", "Telefone", "Plano", "Status", "Vencimento", "Tag RFID"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, s in enumerate(students, 2):
        ws.cell(row=row, column=1, value=s.get("nome", ""))
        ws.cell(row=row, column=2, value=s.get("email", ""))
        ws.cell(row=row, column=3, value=s.get("cpf", ""))
        ws.cell(row=row, column=4, value=s.get("telefone", ""))
        ws.cell(row=row, column=5, value=plans_map.get(s.get("plano_id", ""), "-"))
        ws.cell(row=row, column=6, value=s.get("status", ""))
        venc = s.get("data_vencimento", "")
        try:
            venc = datetime.fromisoformat(venc).strftime("%d/%m/%Y")
        except Exception:
            pass
        ws.cell(row=row, column=7, value=venc)
        ws.cell(row=row, column=8, value=s.get("tag_rfid", ""))

    for col in range(1, 9):
        ws.column_dimensions[chr(64+col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=alunos_gymbro.xlsx"})

@app.get("/api/reports/students/pdf")
async def export_students_pdf(user=Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    students = await db.students.find({}, {"_id": 0}).sort("nome", 1).to_list(5000)
    plans_list = await db.plans.find({}, {"_id": 0}).to_list(100)
    plans_map = {p["plan_id"]: p["nome"] for p in plans_list}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("GymBro - Relatorio de Alunos", styles['Title']))
    elements.append(Paragraph(f"Gerado em: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    data = [["Nome", "Email", "CPF", "Plano", "Status", "Vencimento"]]
    for s in students:
        venc = s.get("data_vencimento", "")
        try:
            venc = datetime.fromisoformat(venc).strftime("%d/%m/%Y")
        except Exception:
            pass
        data.append([s.get("nome",""), s.get("email",""), s.get("cpf",""),
                     plans_map.get(s.get("plano_id",""), "-"), s.get("status","").upper(), venc])

    table = Table(data, colWidths=[120, 150, 100, 80, 60, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1a1a1a")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=alunos_gymbro.pdf"})

@app.get("/api/reports/access-logs/excel")
async def export_access_logs_excel(user=Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    logs = await db.access_logs.find({}, {"_id": 0}).sort("timestamp", -1).limit(5000).to_list(5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Acessos"
    headers = ["Data/Hora", "Aluno", "Tag/ID", "Tipo", "Status", "Motivo"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, l in enumerate(logs, 2):
        ts = l.get("timestamp", "")
        try:
            ts = datetime.fromisoformat(ts).strftime("%d/%m/%Y %H:%M")
        except Exception:
            pass
        ws.cell(row=row, column=1, value=ts)
        ws.cell(row=row, column=2, value=l.get("student_name", ""))
        ws.cell(row=row, column=3, value=l.get("tag_id", ""))
        ws.cell(row=row, column=4, value=l.get("tipo", ""))
        ws.cell(row=row, column=5, value="Liberado" if l.get("autorizado") else "Negado")
        ws.cell(row=row, column=6, value=l.get("motivo", ""))

    for col in range(1, 7):
        ws.column_dimensions[chr(64+col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=acessos_gymbro.xlsx"})

@app.get("/api/reports/financial/excel")
async def export_financial_excel(user=Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    students = await db.students.find({"status": "ativo"}, {"_id": 0}).to_list(5000)
    plans_list = await db.plans.find({}, {"_id": 0}).to_list(100)
    plans_map = {p["plan_id"]: p for p in plans_list}

    wb = Workbook()
    ws = wb.active
    ws.title = "Financeiro"
    headers = ["Aluno", "Email", "Plano", "Valor (R$)", "Vencimento", "Status"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    total = 0
    for row, s in enumerate(students, 2):
        plan = plans_map.get(s.get("plano_id", ""), {})
        valor = plan.get("valor", 0)
        total += valor
        venc = s.get("data_vencimento", "")
        try:
            venc = datetime.fromisoformat(venc).strftime("%d/%m/%Y")
        except Exception:
            pass
        ws.cell(row=row, column=1, value=s.get("nome", ""))
        ws.cell(row=row, column=2, value=s.get("email", ""))
        ws.cell(row=row, column=3, value=plan.get("nome", "-"))
        ws.cell(row=row, column=4, value=valor)
        ws.cell(row=row, column=5, value=venc)
        ws.cell(row=row, column=6, value=s.get("status", "").upper())

    row_total = len(students) + 2
    ws.cell(row=row_total, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row_total, column=4, value=total).font = Font(bold=True)

    for col in range(1, 7):
        ws.column_dimensions[chr(64+col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=financeiro_gymbro.xlsx"})

# ============== CATRACA REMOTE CONTROL ==============

@app.post("/api/catraca/command")
async def catraca_command(data: CatracaCommand, user=Depends(get_current_user)):
    """Queue a command for the local agent to execute on the turnstile"""
    cmd_id = f"cmd_{uuid.uuid4().hex[:12]}"
    doc = {
        "cmd_id": cmd_id, "action": data.action, "message": data.message or "",
        "academy_id": data.academy_id or user.get("academy_id", ""),
        "status": "pending", "created_by": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.catraca_commands.insert_one(doc)
    await ws_manager.broadcast({"type": "catraca_command", "data": {"cmd_id": cmd_id, "action": data.action, "message": data.message}})
    return {k: v for k, v in doc.items() if k != "_id"}

@app.get("/api/catraca/commands")
async def list_catraca_commands(user=Depends(get_current_user), limit: int = 20):
    return await db.catraca_commands.find({}, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)

@app.post("/api/catraca/ilnet2/execute")
async def execute_catraca_ilnet2(data: CatracaLanExecute, user=Depends(get_current_user)):
    academy = await db.academies.find_one({"academy_id": data.academy_id}, {"_id": 0})
    if not academy:
        raise HTTPException(status_code=404, detail="Academia nao encontrada")
    host = academy.get("catraca_ip")
    port = int(academy.get("catraca_port", 0) or 0)
    if not host or port <= 0:
        raise HTTPException(status_code=400, detail="Academia sem IP/porta da catraca configurados")

    payload = ilnet2_payload_for_action(data.action, data.message or "", data.raw_hex or "")
    cmd_id = f"cmd_{uuid.uuid4().hex[:12]}"
    created_at = datetime.now(timezone.utc).isoformat()
    await db.catraca_commands.insert_one({
        "cmd_id": cmd_id,
        "action": data.action,
        "message": data.message,
        "academy_id": data.academy_id,
        "status": "sending",
        "transport": "tcp/lan/ilnet2",
        "target_ip": host,
        "target_port": port,
        "created_at": created_at,
        "created_by": user.get("user_id", ""),
        "raw_hex": payload.hex(),
    })

    response_text = ""
    try:
        response_text = await send_tcp_command(host, port, payload, float(data.timeout_seconds or 3.0))
        await db.catraca_commands.update_one(
            {"cmd_id": cmd_id},
            {"$set": {"status": "executed", "executed_at": datetime.now(timezone.utc).isoformat(), "device_response": response_text[:500]}}
        )
    except HTTPException as e:
        await db.catraca_commands.update_one(
            {"cmd_id": cmd_id},
            {"$set": {"status": "error", "executed_at": datetime.now(timezone.utc).isoformat(), "error": str(e.detail)}}
        )
        raise

    return {"cmd_id": cmd_id, "status": "executed", "device_response": response_text}

@app.get("/api/catraca/pending")
async def get_pending_commands(academy_id: str = ""):
    """Endpoint polled by the local agent to get pending commands"""
    q = {"status": "pending"}
    if academy_id:
        q["academy_id"] = academy_id
    cmds = await db.catraca_commands.find(q, {"_id": 0}).sort("created_at", 1).to_list(50)
    # Mark as delivered
    for cmd in cmds:
        await db.catraca_commands.update_one({"cmd_id": cmd["cmd_id"]}, {"$set": {"status": "delivered"}})
    return cmds

@app.put("/api/catraca/commands/{cmd_id}/status")
async def update_command_status(cmd_id: str, request: Request):
    """Local agent reports command execution status"""
    body = await request.json()
    new_status = body.get("status", "executed")
    await db.catraca_commands.update_one({"cmd_id": cmd_id}, {"$set": {"status": new_status, "executed_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "Status atualizado"}

# ============== WEBSOCKET ==============

@app.websocket("/api/ws")
async def websocket_endpoint(ws: WebSocket):
    await ws_manager.connect(ws)
    try:
        while True:
            data = await ws.receive_text()
            # Keep-alive / ping
    except WebSocketDisconnect:
        ws_manager.disconnect(ws)

# ============== SEED DATA ==============

@app.post("/api/seed")
async def seed_data():
    existing_plans = await db.plans.count_documents({})
    if existing_plans > 0:
        return {"message": "Dados ja existem"}

    # Create default academy
    await db.academies.insert_one({
        "academy_id": "acad_matriz", "nome": "GymBro Matriz",
        "endereco": "Av. Paulista, 1000 - Sao Paulo, SP",
        "telefone": "(11) 99999-9999", "cnpj": "12.345.678/0001-00",
        "email": "matriz@gymbro.com.br", "catraca_ip": "192.168.1.9",
        "catraca_port": 7878, "ativo": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    plans = [
        {"plan_id": "plan_mensal", "nome": "Mensal", "valor": 139.90, "duracao_dias": 30, "descricao": "Acesso completo por 30 dias", "ativo": True, "created_at": datetime.now(timezone.utc).isoformat()},
        {"plan_id": "plan_trimestral", "nome": "Trimestral", "valor": 369.90, "duracao_dias": 90, "descricao": "Acesso completo por 90 dias - Economize 15%", "ativo": True, "created_at": datetime.now(timezone.utc).isoformat()},
        {"plan_id": "plan_semestral", "nome": "Semestral", "valor": 669.90, "duracao_dias": 180, "descricao": "Acesso completo por 180 dias - Economize 20%", "ativo": True, "created_at": datetime.now(timezone.utc).isoformat()},
        {"plan_id": "plan_anual", "nome": "Anual", "valor": 1249.90, "duracao_dias": 365, "descricao": "Acesso completo por 365 dias - Economize 25%", "ativo": True, "created_at": datetime.now(timezone.utc).isoformat()},
    ]
    await db.plans.insert_many(plans)

    future = (datetime.now(timezone.utc) + timedelta(days=25)).isoformat()
    future_3d = (datetime.now(timezone.utc) + timedelta(days=3)).isoformat()
    future_5d = (datetime.now(timezone.utc) + timedelta(days=5)).isoformat()
    past = (datetime.now(timezone.utc) - timedelta(days=5)).isoformat()
    students = [
        {"student_id": "std_001", "nome": "Carlos Silva", "email": "carlos@email.com", "cpf": "123.456.789-00", "telefone": "(11) 99999-0001", "plano_id": "plan_mensal", "tag_rfid": "0000000001", "biometria_id": "", "status": "ativo", "data_vencimento": future, "academy_id": "acad_matriz", "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()},
        {"student_id": "std_002", "nome": "Ana Souza", "email": "ana@email.com", "cpf": "234.567.890-11", "telefone": "(11) 99999-0002", "plano_id": "plan_trimestral", "tag_rfid": "0000000002", "biometria_id": "", "status": "ativo", "data_vencimento": future_3d, "academy_id": "acad_matriz", "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()},
        {"student_id": "std_003", "nome": "Bruno Oliveira", "email": "bruno@email.com", "cpf": "345.678.901-22", "telefone": "(11) 99999-0003", "plano_id": "plan_semestral", "tag_rfid": "0000000003", "biometria_id": "", "status": "ativo", "data_vencimento": future, "academy_id": "acad_matriz", "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()},
        {"student_id": "std_004", "nome": "Fernanda Lima", "email": "fernanda@email.com", "cpf": "456.789.012-33", "telefone": "(11) 99999-0004", "plano_id": "plan_mensal", "tag_rfid": "0000000004", "biometria_id": "", "status": "inativo", "data_vencimento": past, "academy_id": "acad_matriz", "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()},
        {"student_id": "std_005", "nome": "Ricardo Santos", "email": "ricardo@email.com", "cpf": "567.890.123-44", "telefone": "(11) 99999-0005", "plano_id": "plan_anual", "tag_rfid": "0000000005", "biometria_id": "1", "status": "ativo", "data_vencimento": future, "academy_id": "acad_matriz", "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()},
        {"student_id": "std_006", "nome": "Juliana Costa", "email": "juliana@email.com", "cpf": "678.901.234-55", "telefone": "(11) 99999-0006", "plano_id": "plan_trimestral", "tag_rfid": "0000000006", "biometria_id": "", "status": "ativo", "data_vencimento": future_5d, "academy_id": "acad_matriz", "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()},
        {"student_id": "std_007", "nome": "Pedro Almeida", "email": "pedro@email.com", "cpf": "789.012.345-66", "telefone": "(11) 99999-0007", "plano_id": "plan_semestral", "tag_rfid": "0000000007", "biometria_id": "2", "status": "ativo", "data_vencimento": future, "academy_id": "acad_matriz", "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()},
        {"student_id": "std_008", "nome": "Mariana Rocha", "email": "mariana@email.com", "cpf": "890.123.456-77", "telefone": "(11) 99999-0008", "plano_id": "plan_anual", "tag_rfid": "0000000008", "biometria_id": "", "status": "ativo", "data_vencimento": future, "academy_id": "acad_matriz", "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()},
    ]
    await db.students.insert_many(students)

    admin_exists = await db.users.find_one({"email": "admin@gymbro.com"}, {"_id": 0})
    if not admin_exists:
        await db.users.insert_one({
            "user_id": "user_admin001", "email": "admin@gymbro.com",
            "name": "Admin GymBro", "password": pwd_context.hash("admin123"),
            "role": "super_admin", "picture": "", "academy_id": "",
            "created_at": datetime.now(timezone.utc),
        })

    access_logs = []
    names = ["Carlos Silva", "Ana Souza", "Bruno Oliveira", "Fernanda Lima", "Ricardo Santos", "Juliana Costa", "Pedro Almeida", "Mariana Rocha"]
    for i in range(15):
        hours_ago = i * 1.5
        ts = (datetime.now(timezone.utc) - timedelta(hours=hours_ago)).isoformat()
        idx = i % len(names)
        access_logs.append({
            "log_id": f"log_seed_{i}", "tag_id": f"000000000{idx+1}",
            "tipo": "rfid" if i % 3 != 2 else "biometria",
            "student_id": f"std_00{idx+1}", "student_name": names[idx],
            "autorizado": i != 3, "motivo": "Acesso liberado" if i != 3 else "Assinatura inativa",
            "academy_id": "acad_matriz",
            "timestamp": ts,
        })
    await db.access_logs.insert_many(access_logs)

    return {"message": "Dados de teste criados com sucesso"}

# ============== BILLING STATUS ==============

@app.get("/api/academies/{academy_id}/billing")
async def get_academy_billing(academy_id: str, user=Depends(get_current_user)):
    """Get billing status and payment history for an academy."""
    academy = await db.academies.find_one({"academy_id": academy_id}, {"_id": 0})
    if not academy:
        raise HTTPException(status_code=404, detail="Academia nao encontrada")
    
    # Get latest billing records (up to last 12 months)
    billing_records = await db.academy_billing.find(
        {"academy_id": academy_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(12).to_list(12)
    
    return {
        "academy_id": academy_id,
        "academy_name": academy.get("nome", ""),
        "status": academy.get("billing_status", "trial"),
        "trial_until": academy.get("trial_until"),
        "paid_until": academy.get("paid_until"),
        "billing_history": billing_records,
    }

# ============== STUDENT PROGRESS (EVOLUÇÃO) ==============

@app.post("/api/students/{student_id}/progress")
async def record_student_progress(student_id: str, request: Request, user=Depends(get_current_user)):
    """Record student progress (weight, height, measurements, notes, photos)."""
    body = await request.json()
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    
    progress_id = f"prog_{uuid.uuid4().hex[:12]}"
    progress_doc = {
        "progress_id": progress_id,
        "student_id": student_id,
        "academy_id": student.get("academy_id", ""),
        "date": body.get("date") or datetime.now(timezone.utc).isoformat(),
        "weight_kg": body.get("weight_kg"),
        "height_cm": body.get("height_cm"),
        "chest_cm": body.get("chest_cm"),
        "waist_cm": body.get("waist_cm"),
        "hip_cm": body.get("hip_cm"),
        "notes": body.get("notes", ""),
        "photos": body.get("photos", []),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.student_progress.insert_one(progress_doc)
    
    # Update student's latest measurements
    await db.students.update_one(
        {"student_id": student_id},
        {"$set": {
            "peso": body.get("weight_kg") or student.get("peso"),
            "altura": body.get("height_cm") or student.get("altura"),
        }}
    )
    
    return await db.student_progress.find_one({"progress_id": progress_id}, {"_id": 0})

@app.get("/api/students/{student_id}/progress")
async def list_student_progress(student_id: str, user=Depends(get_current_user), limit: int = 50):
    """Get student progress history (evolução)."""
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    
    progress_records = await db.student_progress.find(
        {"student_id": student_id},
        {"_id": 0}
    ).sort("date", -1).limit(limit).to_list(limit)
    
    return {
        "student_id": student_id,
        "student_name": student.get("nome"),
        "progress_records": progress_records,
    }

# ============== ATTENDANCE (PRESENÇA) ==============

@app.post("/api/students/{student_id}/attendance")
async def record_attendance(student_id: str, request: Request, user=Depends(get_current_user)):
    """Record student attendance (presença)."""
    body = await request.json()
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    
    attendance_id = f"att_{uuid.uuid4().hex[:12]}"
    attendance_doc = {
        "attendance_id": attendance_id,
        "student_id": student_id,
        "academy_id": student.get("academy_id", ""),
        "date_time": body.get("date_time") or datetime.now(timezone.utc).isoformat(),
        "method": body.get("method", "manual"),  # manual, qr, webauthn, rfid
        "gate": body.get("gate", "default"),
        "notes": body.get("notes", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.attendance.insert_one(attendance_doc)
    return await db.attendance.find_one({"attendance_id": attendance_id}, {"_id": 0})

@app.get("/api/students/{student_id}/attendance")
async def list_student_attendance(student_id: str, user=Depends(get_current_user), limit: int = 100):
    """Get student attendance history (presença)."""
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Aluno nao encontrado")
    
    attendance_records = await db.attendance.find(
        {"student_id": student_id},
        {"_id": 0}
    ).sort("date_time", -1).limit(limit).to_list(limit)
    
    return {
        "student_id": student_id,
        "student_name": student.get("nome"),
        "attendance_records": attendance_records,
        "total_presencas_este_mes": len([a for a in attendance_records if a.get("date_time", "").startswith(datetime.now(timezone.utc).strftime("%Y-%m"))]),
    }

@app.get("/api/health")
async def health():
    return {"status": "ok", "service": "GymBro API", "version": "2.0.0"}
