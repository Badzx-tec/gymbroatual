import secrets
from datetime import datetime
from io import BytesIO

from fastapi import (
    APIRouter,
    Depends,
    Header,
    HTTPException,
    Request,
    WebSocket,
    WebSocketDisconnect,
)
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from app.core.deps import get_current_actor, require_active_subscription, require_roles
from app.core.time import UTC
from app.db.mongo import get_db

from . import billing as billing_routes
from . import gyms as gym_routes
from . import turnstiles as turnstile_routes

router = APIRouter()


def _clean_doc(doc: dict) -> dict:
    sanitized = dict(doc)
    sanitized.pop("_id", None)
    return sanitized


def _as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo:
            value = value.astimezone(UTC)
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, bool):
        return "Sim" if value else "Nao"
    return str(value)


def _xlsx_response(filename: str, sheet_name: str, headers: list[str], rows: list[list[str]]):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    for index, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(14, len(header) + 2)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _students_pdf_response(filename: str, rows: list[list[str]]):
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    y = height - 36

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(32, y, "Relatorio de Alunos - GymBro")
    y -= 18
    pdf.setFont("Helvetica", 9)
    pdf.drawString(32, y, f"Gerado em: {datetime.now(UTC).strftime('%d/%m/%Y %H:%M UTC')}")
    y -= 20

    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(32, y, "Nome")
    pdf.drawString(210, y, "Email")
    pdf.drawString(390, y, "CPF")
    pdf.drawString(480, y, "Status")
    y -= 12
    pdf.setFont("Helvetica", 8)

    for row in rows:
        if y < 48:
            pdf.showPage()
            y = height - 36
            pdf.setFont("Helvetica-Bold", 8)
            pdf.drawString(32, y, "Nome")
            pdf.drawString(210, y, "Email")
            pdf.drawString(390, y, "CPF")
            pdf.drawString(480, y, "Status")
            y -= 12
            pdf.setFont("Helvetica", 8)

        nome = (row[1] or "")[:38]
        email = (row[2] or "")[:33]
        cpf = (row[3] or "")[:18]
        status = (row[6] or "")[:10]
        pdf.drawString(32, y, nome)
        pdf.drawString(210, y, email)
        pdf.drawString(390, y, cpf)
        pdf.drawString(480, y, status)
        y -= 12

    pdf.save()
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/auth/logout")
async def logout() -> dict:
    return {"message": "ok"}


@router.get("/dashboard")
async def dashboard(actor: dict = Depends(require_active_subscription)):
    return await gym_routes.dashboard(actor)


@router.get("/dashboard/charts")
async def dashboard_charts(actor: dict = Depends(require_active_subscription)) -> dict:
    return await gym_routes.dashboard_charts(actor)


@router.get("/access-logs")
async def access_logs(limit: int = 100, actor: dict = Depends(require_active_subscription)):
    return await gym_routes.access_logs(limit=limit, owner=actor)


@router.get("/plans")
async def plans(actor: dict = Depends(require_active_subscription)):
    return await gym_routes.list_plans(actor)


@router.get("/plans/public")
async def plans_public():
    return await gym_routes.list_plans_public()


@router.post("/plans")
async def create_plan(payload: dict, actor: dict = Depends(require_active_subscription)):
    return await gym_routes.create_plan(payload, actor)


@router.put("/plans/{plan_id}")
async def update_plan(
    plan_id: str, payload: dict, actor: dict = Depends(require_active_subscription)
):
    return await gym_routes.update_plan(plan_id, payload, actor)


@router.delete("/plans/{plan_id}")
async def delete_plan(plan_id: str, actor: dict = Depends(require_active_subscription)):
    return await gym_routes.delete_plan(plan_id, actor)


@router.get("/academies")
async def list_academies(actor: dict = Depends(require_active_subscription)):
    return await gym_routes.list_gyms(actor)


@router.post("/academies")
async def create_academy(payload: dict, actor: dict = Depends(require_active_subscription)):
    raise HTTPException(status_code=403, detail="Gestao de franquias desabilitada")


@router.put("/academies/{academy_id}")
async def update_academy(
    academy_id: str, payload: dict, actor: dict = Depends(require_active_subscription)
):
    raise HTTPException(status_code=403, detail="Gestao de franquias desabilitada")


@router.delete("/academies/{academy_id}")
async def delete_academy(academy_id: str, actor: dict = Depends(require_active_subscription)):
    raise HTTPException(status_code=403, detail="Gestao de franquias desabilitada")


@router.get("/academies/{academy_id}/stats")
async def academy_stats(academy_id: str, actor: dict = Depends(require_active_subscription)):
    db = get_db()
    q = {"owner_id": actor["owner_id"], "gym_id": academy_id}
    total = await db.students.count_documents(q)
    active = await db.students.count_documents({**q, "status": "ativo"})
    return {"total_alunos": total, "alunos_ativos": active, "faturamento": 0}


@router.get("/academies/{academy_id}/billing")
async def academy_billing(
    academy_id: str, actor: dict = Depends(require_roles("OWNER", "MANAGER"))
):
    status = await billing_routes.subscription_status(actor)
    return {
        "academy_id": academy_id,
        "academy_name": "Academia",
        "status": status.status,
        "trial_until": status.trial_ends_at,
        "paid_until": status.current_period_end,
        "billing_history": [],
    }


@router.post("/payments/academy/subscription/checkout")
async def academy_checkout(_: dict, actor: dict = Depends(require_roles("OWNER", "MANAGER"))):
    return await billing_routes.subscription_checkout(actor)


@router.get("/notifications")
async def notifications(limit: int = 50, actor: dict = Depends(require_active_subscription)):
    db = get_db()
    return (
        await db.notifications.find({"owner_id": actor["owner_id"]}, {"_id": 0})
        .sort("created_at", -1)
        .limit(limit)
        .to_list(limit)
    )


@router.post("/notifications/check-expiring")
async def notifications_check(actor: dict = Depends(require_active_subscription)):
    db = get_db()
    count = 0
    for student in await db.students.find(
        {"owner_id": actor["owner_id"], "status": "ativo"}, {"_id": 0}
    ).to_list(500):
        if not student.get("email"):
            continue
        notif = {
            "notif_id": f"n_{secrets.token_hex(6)}",
            "owner_id": actor["owner_id"],
            "student_id": student["student_id"],
            "student_email": student.get("email", ""),
            "tipo": "vencimento",
            "titulo": "Acompanhar aluno",
            "mensagem": f"Aluno {student.get('nome', '')} sem treino atualizado recentemente.",
            "lida": False,
            "email_enviado": False,
            "email_status": "simulado",
            "created_at": datetime.now(UTC),
        }
        await db.notifications.insert_one(notif)
        count += 1
    return {"message": f"{count} notificacoes geradas"}


@router.put("/notifications/{notif_id}/read")
async def notification_read(notif_id: str, actor: dict = Depends(require_active_subscription)):
    db = get_db()
    await db.notifications.update_one(
        {"notif_id": notif_id, "owner_id": actor["owner_id"]}, {"$set": {"lida": True}}
    )
    return {"message": "ok"}


@router.delete("/notifications/{notif_id}")
async def notification_delete(notif_id: str, actor: dict = Depends(require_active_subscription)):
    db = get_db()
    await db.notifications.delete_one({"notif_id": notif_id, "owner_id": actor["owner_id"]})
    return {"message": "ok"}


@router.post("/catraca/command")
async def catraca_command(payload: dict, actor: dict = Depends(require_active_subscription)):
    db = get_db()
    doc = {
        "cmd_id": f"cmd_{secrets.token_hex(6)}",
        "owner_id": actor["owner_id"],
        "action": payload.get("action"),
        "message": payload.get("message", ""),
        "status": "pending",
        "created_at": datetime.now(UTC),
    }
    await db.catraca_commands.insert_one(doc)
    return _clean_doc(doc)


@router.get("/catraca/commands")
async def catraca_commands(actor: dict = Depends(require_active_subscription)):
    db = get_db()
    return (
        await db.catraca_commands.find({"owner_id": actor["owner_id"]}, {"_id": 0})
        .sort("created_at", -1)
        .limit(50)
        .to_list(50)
    )


@router.get("/webhook-logs")
async def webhook_logs(limit: int = 50, actor: dict = Depends(require_active_subscription)):
    db = get_db()
    return (
        await db.billing_events.find({}, {"_id": 0})
        .sort("received_at", -1)
        .limit(limit)
        .to_list(limit)
    )


@router.get("/reports/students/excel")
async def export_students_excel(actor: dict = Depends(require_active_subscription)):
    db = get_db()
    students = (
        await db.students.find({"owner_id": actor["owner_id"]}, {"_id": 0})
        .sort("created_at", -1)
        .to_list(5000)
    )
    headers = [
        "ID",
        "Nome",
        "Email",
        "CPF",
        "Telefone",
        "Plano",
        "Status",
        "Vencimento",
        "Criado em",
    ]
    rows = [
        [
            _as_text(student.get("student_id")),
            _as_text(student.get("nome")),
            _as_text(student.get("email")),
            _as_text(student.get("cpf")),
            _as_text(student.get("telefone")),
            _as_text(student.get("plano_id")),
            _as_text(student.get("status")),
            _as_text(student.get("data_vencimento")),
            _as_text(student.get("created_at")),
        ]
        for student in students
    ]
    return _xlsx_response("alunos_gymbro.xlsx", "Alunos", headers, rows)


@router.get("/reports/students/pdf")
async def export_students_pdf(actor: dict = Depends(require_active_subscription)):
    db = get_db()
    students = (
        await db.students.find({"owner_id": actor["owner_id"]}, {"_id": 0})
        .sort("created_at", -1)
        .to_list(3000)
    )
    rows = [
        [
            _as_text(student.get("student_id")),
            _as_text(student.get("nome")),
            _as_text(student.get("email")),
            _as_text(student.get("cpf")),
            _as_text(student.get("telefone")),
            _as_text(student.get("plano_id")),
            _as_text(student.get("status")),
            _as_text(student.get("created_at")),
        ]
        for student in students
    ]
    return _students_pdf_response("alunos_gymbro.pdf", rows)


@router.get("/reports/access-logs/excel")
async def export_access_excel(actor: dict = Depends(require_active_subscription)):
    db = get_db()
    logs = (
        await db.access_logs.find({"owner_id": actor["owner_id"]}, {"_id": 0})
        .sort("created_at", -1)
        .to_list(10000)
    )
    headers = [
        "Data/Hora",
        "Tipo",
        "Aluno",
        "Funcionario",
        "Metodo",
        "Autorizado",
        "Motivo",
    ]
    rows = [
        [
            _as_text(log.get("timestamp") or log.get("created_at")),
            _as_text(log.get("subject_type") or log.get("tipo")),
            _as_text(log.get("student_name")),
            _as_text(log.get("employee_name")),
            _as_text(log.get("method")),
            _as_text(log.get("autorizado")),
            _as_text(log.get("motivo") or log.get("reason")),
        ]
        for log in logs
    ]
    return _xlsx_response("acessos_gymbro.xlsx", "Acessos", headers, rows)


@router.get("/reports/financial/excel")
async def export_financial_excel(actor: dict = Depends(require_roles("OWNER", "MANAGER"))):
    db = get_db()
    invoices = (
        await db.invoices.find({"owner_id": actor["owner_id"]}, {"_id": 0})
        .sort("created_at", -1)
        .to_list(5000)
    )
    charges = (
        await db.student_charges.find({"owner_id": actor["owner_id"]}, {"_id": 0})
        .sort("created_at", -1)
        .to_list(5000)
    )
    headers = ["Tipo", "Referencia", "Status", "Valor", "Vencimento", "Pago em", "Origem"]
    rows: list[list[str]] = []
    for invoice in invoices:
        rows.append(
            [
                "Assinatura SaaS",
                _as_text(invoice.get("period_label") or invoice.get("invoice_id")),
                _as_text(invoice.get("status")),
                _as_text(invoice.get("amount")),
                _as_text(invoice.get("due_at") or invoice.get("created_at")),
                _as_text(invoice.get("paid_at")),
                "billing.invoices",
            ]
        )
    for charge in charges:
        rows.append(
            [
                "Contrato Aluno",
                _as_text(charge.get("charge_id")),
                _as_text(charge.get("status")),
                _as_text(charge.get("amount")),
                _as_text(charge.get("due_at")),
                _as_text(charge.get("paid_at")),
                "student_charges",
            ]
        )
    return _xlsx_response("financeiro_gymbro.xlsx", "Financeiro", headers, rows)


@router.post("/students/{student_id}/biometria")
async def students_biometria(
    student_id: str, payload: dict, actor: dict = Depends(require_active_subscription)
):
    db = get_db()
    await db.students.update_one(
        {"student_id": student_id, "owner_id": actor["owner_id"]},
        {"$set": {"biometria_id": payload.get("biometria_id")}},
    )
    return {"message": "ok"}


@router.post("/students/{student_id}/progress")
async def student_progress(
    student_id: str, payload: dict, actor: dict = Depends(require_active_subscription)
):
    db = get_db()
    doc = {
        "progress_id": f"prg_{secrets.token_hex(6)}",
        "student_id": student_id,
        "owner_id": actor["owner_id"],
        **payload,
        "created_at": datetime.now(UTC),
    }
    await db.measurements.insert_one(doc)
    return _clean_doc(doc)


@router.get("/students/{student_id}/progress")
async def student_progress_list(
    student_id: str, actor: dict = Depends(require_active_subscription), limit: int = 50
):
    db = get_db()
    records = (
        await db.measurements.find(
            {"student_id": student_id, "owner_id": actor["owner_id"]}, {"_id": 0}
        )
        .sort("created_at", -1)
        .limit(limit)
        .to_list(limit)
    )
    return {"student_id": student_id, "progress_records": records}


@router.post("/students/{student_id}/passkey/register")
async def register_passkey(
    student_id: str, _: dict, actor: dict = Depends(require_active_subscription)
):
    return {
        "message": "Passkey placeholder",
        "student_id": student_id,
        "owner_id": actor["owner_id"],
    }


@router.post("/students/{student_id}/passkey/register/options")
async def passkey_opts(student_id: str, actor: dict = Depends(require_active_subscription)):
    return {
        "state_id": f"st_{student_id}",
        "publicKey": {"challenge": "", "user": {"id": ""}, "excludeCredentials": []},
    }


@router.post("/students/{student_id}/passkey/register/verify")
async def passkey_verify(
    student_id: str, _: dict, actor: dict = Depends(require_active_subscription)
):
    return {"message": "ok", "student_id": student_id, "owner_id": actor["owner_id"]}


@router.get("/students/{student_id}/passkeys")
async def list_passkeys(student_id: str, actor: dict = Depends(require_active_subscription)):
    return {"student_id": student_id, "webauthn_credentials": []}


@router.post("/seed")
async def seed(actor: dict = Depends(get_current_actor)):
    db = get_db()
    now = datetime.now(UTC)
    plans = await db.plans.count_documents({"owner_id": actor["owner_id"]})
    if plans == 0:
        await db.plans.insert_many(
            [
                {
                    "plan_id": "mensal",
                    "owner_id": actor["owner_id"],
                    "nome": "Mensal",
                    "valor": 139.9,
                    "duracao_dias": 30,
                    "ativo": True,
                    "created_at": now,
                },
                {
                    "plan_id": "trimestral",
                    "owner_id": actor["owner_id"],
                    "nome": "Trimestral",
                    "valor": 369.9,
                    "duracao_dias": 90,
                    "ativo": True,
                    "created_at": now,
                },
            ]
        )
    return {"message": "ok"}


@router.websocket("/ws")
async def ws_endpoint(websocket: WebSocket):
    await websocket.accept()
    try:
        while True:
            await websocket.receive_text()
            await websocket.send_json({"type": "ping"})
    except WebSocketDisconnect:
        return


@router.post("/turnstile/decision")
async def turnstile_decision_alias(
    payload: dict,
    request: Request,
    x_device_token: str | None = Header(default=None),
):
    return await turnstile_routes.turnstile_decision(
        payload=payload,
        request=request,
        x_device_token=x_device_token,
    )


@router.post("/turnstile/events")
async def turnstile_events_alias(
    payload: dict,
    request: Request,
    x_device_token: str | None = Header(default=None),
):
    return await turnstile_routes.turnstile_event(
        payload=payload,
        request=request,
        x_device_token=x_device_token,
    )
