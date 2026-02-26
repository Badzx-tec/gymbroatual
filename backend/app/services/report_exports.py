from datetime import datetime
from io import BytesIO

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from app.core.time import UTC


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo:
            value = value.astimezone(UTC)
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, bool):
        return "Sim" if value else "Nao"
    return str(value)


def xlsx_response(filename: str, sheet_name: str, headers: list[str], rows: list[list[str]]):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    for index, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(14, len(header) + 2)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def students_pdf_response(filename: str, rows: list[list[str]]):
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    y = height - 36

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(32, y, "Relatorio de Alunos - GymBro")
    y -= 18
    pdf.setFont("Helvetica", 9)
    pdf.drawString(32, y, f"Gerado em: {datetime.now(UTC).strftime('%d/%m/%Y %H:%M UTC')}")
    y -= 20

    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(32, y, "Nome")
    pdf.drawString(210, y, "Email")
    pdf.drawString(390, y, "CPF")
    pdf.drawString(480, y, "Status")
    y -= 12
    pdf.setFont("Helvetica", 8)

    for row in rows:
        if y < 48:
            pdf.showPage()
            y = height - 36
            pdf.setFont("Helvetica-Bold", 8)
            pdf.drawString(32, y, "Nome")
            pdf.drawString(210, y, "Email")
            pdf.drawString(390, y, "CPF")
            pdf.drawString(480, y, "Status")
            y -= 12
            pdf.setFont("Helvetica", 8)

        nome = (row[1] or "")[:38]
        email = (row[2] or "")[:33]
        cpf = (row[3] or "")[:18]
        status = (row[6] or "")[:10]
        pdf.drawString(32, y, nome)
        pdf.drawString(210, y, email)
        pdf.drawString(390, y, cpf)
        pdf.drawString(480, y, status)
        y -= 12

    pdf.save()
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
